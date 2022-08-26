from sys import exit
from os import getcwd
from os import path
from os import makedirs
from openpyxl import load_workbook as lw
import qrcode as qr


def generateFolder(folderName):

    newcwd = getCurrentDir()

    newpath = newcwd + "\\\\" + folderName + "\\\\"

    if not path.exists(newpath):
        makedirs(newpath)

    return newpath

def getCurrentDir():
    cwd = str(getcwd())
    newcwd = ""

    for charac in range(len(cwd)):
        if (cwd[charac] != str("\\")):
            newcwd += cwd[charac]
        else:
            newcwd += "\\\\"

    return newcwd


def generateQRCode(commandKey, folderName):

        file_path = "QRBoard.xlsx"

        wb = lw(file_path)
        ws = wb.active

        #Check if the cell i- not empty and count
        counter = 1
        embedData = {}

        try:

            while True:

                if ((ws['A'+str(counter+1)].value) != None):
                    counter += 1
                    
                    embedData["Name"] = ws['A'+str(counter+1)].value
                    embedData["LRN"] = ws['B'+str(counter+1)].value
                    embedData["Address"] = ws['C'+str(counter+1)].value
                    embedData["Parent"] = ws['D'+str(counter+1)].value
                    embedData["PhoneNumber"] = ws['E'+str(counter+1)].value

                    img = qr.make(str(embedData))
                    type(img)

                    if (ws['A'+str(counter+1)].value != None):
                        img.save(generateFolder(folderName) + ws['A'+str(counter+1)].value + ".png")
                    else:
                        print("Encountered Blank Cell. Please make sure, the row is not empty")
                        break

                if (counter % 2):
                    print("Generating..")
                elif (counter % 3):
                    print("Generating...")
                else:
                    print("Generating.")
                
            return "QRCode Generated Successfully! It is in Folder Name: " + folderName + "\n"
        
        except:
            return "Unexpected Error Occured. Closing the app automatically"
            sys.exit()
        

def main():

    prnt = ["Direction for use",
    "1. Open the QRBoard.xlsx excel file and follow the template/format. Else, this will not work",
    "2. Copy paste your student details, if a particular property has no details, please put 'NONE' or 'N/A'",
    "3. Start the app by clicking S",
    "4. A prompt that will make a new folder for your generated qr code will appear. Please input your desired filename.",
    "5. For new QR Generation, please exit the app by clicking Q, put the new details in QRBoard.xlsx and open the app",
    "6. Press Q to exit"
    ]

    for prnts in range(len(prnt)):
        print(prnt[prnts] + "\n")

    while True:
        try:
            puts = input("Please press S + Enter to start generating and Q + enter to exit. \n")
            put = puts.lower()
            if ((put) == "s"):
                folderName = input("Please input below the name of the folder you want your QR Code to be generated. Note, the folder must be in the folder containing the app.\n")
                print(generateQRCode(put, folderName))
            elif ((put) == "q"):
                break
            else:
                print(put)
                print("Wrong command key")

        except:
            print("A generic error encountered, app will automatically shutdown. Please contact developer. Soriano")
            sys.exit()

main()




    
    



    









